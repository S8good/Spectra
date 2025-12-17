# nanosense/tools/lspr_export.py
"""
LSPR 仿真数据导出工具
支持 CSV, JSON, PNG, Excel 等多种格式
"""

import json
import csv
from pathlib import Path
from typing import Dict, Any, Tuple
import numpy as np
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pyqtgraph as pg
    from PyQt5.QtGui import QPixmap
    HAS_PYQT = True
except ImportError:
    HAS_PYQT = False


class LSPRDataExporter:
    """LSPR 仿真数据导出器"""
    
    def __init__(self):
        self.export_timestamp = datetime.now().isoformat()
    
    def export_to_csv(self, filepath: str, shift_matrix: np.ndarray, 
                      parameters: Dict[str, Any]) -> bool:
        """
        导出热力图数据为 CSV 格式
        
        Args:
            filepath: 输出文件路径
            shift_matrix: 波长偏移矩阵 (15x15)
            parameters: 仿真参数字典
            
        Returns:
            是否成功导出
        """
        try:
            with open(filepath, 'w', newline='') as f:
                writer = csv.writer(f)
                
                # 写入参数部分
                writer.writerow(['LSPR Sensor Simulation - CSV Export'])
                writer.writerow(['Generated', self.export_timestamp])
                writer.writerow([])
                
                # 写入参数
                writer.writerow(['Parameters:'])
                for key, value in parameters.items():
                    writer.writerow([f"  {key}", value])
                writer.writerow([])
                
                # 写入矩阵数据
                writer.writerow(['Wavelength Shift Matrix (nm):'])
                writer.writerow(['Row/Col'] + [f'Col {i}' for i in range(shift_matrix.shape[1])])
                
                for i, row in enumerate(shift_matrix):
                    writer.writerow([f'Row {i}'] + [f'{val:.4f}' for val in row])
                
                # 写入统计信息
                writer.writerow([])
                writer.writerow(['Statistics:'])
                writer.writerow(['Min', f'{np.min(shift_matrix):.4f}'])
                writer.writerow(['Max', f'{np.max(shift_matrix):.4f}'])
                writer.writerow(['Mean', f'{np.mean(shift_matrix):.4f}'])
                writer.writerow(['Std', f'{np.std(shift_matrix):.4f}'])
            
            return True
        except Exception as e:
            print(f"CSV 导出失败: {str(e)}")
            return False
    
    def export_to_json(self, filepath: str, shift_matrix: np.ndarray,
                       parameters: Dict[str, Any], spectrum_data: Dict = None) -> bool:
        """
        导出仿真数据为 JSON 格式
        
        Args:
            filepath: 输出文件路径
            shift_matrix: 波长偏移矩阵
            parameters: 仿真参数
            spectrum_data: 可选的光谱数据
            
        Returns:
            是否成功导出
        """
        try:
            export_data = {
                'metadata': {
                    'timestamp': self.export_timestamp,
                    'version': '1.0',
                    'format': 'LSPR Simulation Export'
                },
                'parameters': {k: str(v) for k, v in parameters.items()},
                'shift_matrix': shift_matrix.tolist(),
                'statistics': {
                    'min': float(np.min(shift_matrix)),
                    'max': float(np.max(shift_matrix)),
                    'mean': float(np.mean(shift_matrix)),
                    'std': float(np.std(shift_matrix))
                }
            }
            
            if spectrum_data:
                export_data['spectrum'] = spectrum_data
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            return True
        except Exception as e:
            print(f"JSON 导出失败: {str(e)}")
            return False
    
    def export_to_excel(self, filepath: str, shift_matrix: np.ndarray,
                       parameters: Dict[str, Any]) -> bool:
        """
        导出仿真数据为 Excel 格式
        
        Args:
            filepath: 输出文件路径
            shift_matrix: 波长偏移矩阵
            parameters: 仿真参数
            
        Returns:
            是否成功导出
        """
        if not HAS_OPENPYXL:
            print("需要安装 openpyxl 库: pip install openpyxl")
            return False
        
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "LSPR Data"
            
            # 写入标题和时间戳
            ws['A1'] = "LSPR Sensor Simulation Export"
            ws['A2'] = f"Generated: {self.export_timestamp}"
            
            # 写入参数
            row = 4
            ws[f'A{row}'] = "Parameters:"
            row += 1
            for key, value in parameters.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
            
            # 写入矩阵数据
            row += 1
            ws[f'A{row}'] = "Wavelength Shift Matrix (nm):"
            row += 1
            
            # 添加列标题
            ws[f'A{row}'] = "Row/Col"
            for col_idx in range(shift_matrix.shape[1]):
                ws.cell(row=row, column=col_idx+2, value=f"Col {col_idx}")
            row += 1
            
            # 添加数据
            for i, data_row in enumerate(shift_matrix):
                ws.cell(row=row, column=1, value=f"Row {i}")
                for j, val in enumerate(data_row):
                    ws.cell(row=row, column=j+2, value=round(float(val), 4))
                row += 1
            
            # 写入统计信息
            row += 1
            ws[f'A{row}'] = "Statistics:"
            row += 1
            stats = [
                ("Min", float(np.min(shift_matrix))),
                ("Max", float(np.max(shift_matrix))),
                ("Mean", float(np.mean(shift_matrix))),
                ("Std", float(np.std(shift_matrix)))
            ]
            for stat_name, stat_val in stats:
                ws[f'A{row}'] = stat_name
                ws[f'B{row}'] = round(stat_val, 4)
                row += 1
            
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Excel 导出失败: {str(e)}")
            return False
    
    def export_heatmap_to_png(self, filepath: str, shift_matrix: np.ndarray,
                             colormap_name: str = 'turbo') -> bool:
        """
        导出热力图为 PNG 图像
        
        Args:
            filepath: 输出文件路径
            shift_matrix: 波长偏移矩阵
            colormap_name: 色彩映射名称
            
        Returns:
            是否成功导出
        """
        if not HAS_PYQT:
            print("需要 PyQt5 和 PyQtGraph")
            return False
        
        try:
            # 创建图像视图
            img_view = pg.ImageView()
            img_view.setImage(shift_matrix, autoRange=True, autoLevels=True)
            img_view.setColorMap(pg.colormap.get(colormap_name))
            
            # 导出为 PNG
            img_view.export(filepath, toBytes=False, copy=False)
            return True
        except Exception as e:
            print(f"PNG 导出失败: {str(e)}")
            return False
    
    def export_spectrum_to_csv(self, filepath: str, wavelengths: np.ndarray,
                              baseline: np.ndarray, signal: np.ndarray,
                              position: Tuple[int, int]) -> bool:
        """
        导出光谱数据为 CSV 格式
        
        Args:
            filepath: 输出文件路径
            wavelengths: 波长数组
            baseline: 基线数据
            signal: 信号数据
            position: 采集位置 (row, col)
            
        Returns:
            是否成功导出
        """
        try:
            with open(filepath, 'w', newline='') as f:
                writer = csv.writer(f)
                
                writer.writerow(['LSPR Spectrum Data Export'])
                writer.writerow(['Generated', self.export_timestamp])
                writer.writerow(['Position', f'({position[0]}, {position[1]})'])
                writer.writerow([])
                
                writer.writerow(['Wavelength (nm)', 'Baseline', 'Signal'])
                for wl, bl, sig in zip(wavelengths, baseline, signal):
                    writer.writerow([f'{wl:.2f}', f'{bl:.6f}', f'{sig:.6f}'])
            
            return True
        except Exception as e:
            print(f"光谱 CSV 导出失败: {str(e)}")
            return False


def get_supported_formats() -> list:
    """获取支持的导出格式列表"""
    formats = ['CSV', 'JSON', 'PNG']
    if HAS_OPENPYXL:
        formats.append('Excel')
    return formats
